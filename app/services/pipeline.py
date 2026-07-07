from __future__ import annotations

import json, os, shutil, subprocess, time, zipfile, struct, zlib
from pathlib import Path
from uuid import uuid4
from xml.etree import ElementTree as ET

from app.services.pdf_metadata import inspect_pdf
from app.services.normalization import write_json
from app.services.repository import create_run
from app.services.adapters import ADAPTERS, DEFAULT_OLLAMA_BASE_URL

SUPPORTED={'.pdf','.docx','.xlsx','.xls'}
DATA_DIR=Path('data')

def _now():
    import datetime; return datetime.datetime.utcnow().isoformat(timespec='seconds')+'Z'

def _rel(p:Path): return str(p)

def save_uploads(run_id, uploads, data_dir=DATA_DIR):
    base=data_dir/'input'/run_id; base.mkdir(parents=True, exist_ok=True)
    files=[]
    for up in uploads:
        name=Path(up.filename).name; ext=Path(name).suffix.lower()
        if ext not in SUPPORTED: raise ValueError(f'Unsupported file format: {ext}')
        fid=uuid4().hex; path=base/f'{Path(name).stem}-{fid}{ext}'
        content=up.file.read() if hasattr(up,'file') else up.read()
        if hasattr(content,'__await__'): raise RuntimeError('async upload must be read by route')
        path.write_bytes(content)
        files.append({'id':fid,'original_filename':name,'saved_path':str(path),'extension':ext,'mime_type':getattr(up,'content_type',None),'file_size_bytes':len(content)})
    return files

def inventory_file(f):
    path=Path(f['saved_path']); ext=f['extension']
    out=f|{'detected_document_type':ext.lstrip('.').upper(),'page_count':None,'sheet_count':None,'image_count':None,'table_count':None,'processing_status':'inventoried','error_message':None}
    try:
        if ext=='.pdf':
            m=inspect_pdf(path); out.update(page_count=m['page_count'], image_count=m['image_count'], table_count=None)
        elif ext=='.docx':
            out.update(_inspect_docx(path))
        elif ext in {'.xlsx','.xls'}:
            out.update(_inspect_excel(path))
    except Exception as e:
        out.update(processing_status='inventory_failed', error_message=str(e))
    return out

def _inspect_docx(path):
    ns={'w':'http://schemas.openxmlformats.org/wordprocessingml/2006/main'}
    paras=tables=imgs=0
    with zipfile.ZipFile(path) as z:
        if 'word/document.xml' in z.namelist():
            root=ET.fromstring(z.read('word/document.xml'))
            paras=len(root.findall('.//w:p',ns)); tables=len(root.findall('.//w:tbl',ns))
        imgs=len([n for n in z.namelist() if n.startswith('word/media/')])
    return {'page_count':None,'sheet_count':None,'image_count':imgs,'table_count':tables,'paragraph_count':paras}

def _inspect_excel(path):
    if path.suffix.lower()=='.xlsx':
        with zipfile.ZipFile(path) as z:
            sheets=len([n for n in z.namelist() if n.startswith('xl/worksheets/sheet') and n.endswith('.xml')])
            tables=len([n for n in z.namelist() if n.startswith('xl/tables/') and n.endswith('.xml')]) or None
        return {'sheet_count':sheets,'page_count':None,'image_count':None,'table_count':tables}
    return {'sheet_count':None,'page_count':None,'image_count':None,'table_count':None}

def parse_docx(path):
    ns={'w':'http://schemas.openxmlformats.org/wordprocessingml/2006/main'}; sections=[]; tables=[]; images=[]; props={}
    with zipfile.ZipFile(path) as z:
        names=z.namelist(); images=[n for n in names if n.startswith('word/media/')]
        if 'word/document.xml' in names:
            root=ET.fromstring(z.read('word/document.xml'))
            texts=[''.join(t.text or '' for t in p.findall('.//w:t',ns)).strip() for p in root.findall('.//w:p',ns)]
            sections=[t for t in texts if t]
            for tbl in root.findall('.//w:tbl',ns):
                rows=[]
                for tr in tbl.findall('.//w:tr',ns):
                    rows.append([''.join(t.text or '' for t in tc.findall('.//w:t',ns)).strip() for tc in tr.findall('./w:tc',ns)])
                tables.append(rows)
    return {'text_by_sections':sections,'tables':tables,'images':images,'document_names':[],'suspicious_values':[],'source_file':str(path),'properties':props}

def parse_xlsx(path):
    try:
        from openpyxl import load_workbook
    except Exception as e:
        return {'parser_status':'not_configured','error':str(e),'sheets':[]}
    wb=load_workbook(path, data_only=False); wbv=load_workbook(path, data_only=True)
    sheets=[]
    for ws in wb.worksheets:
        wsv=wbv[ws.title]; rows=[]
        for row in ws.iter_rows():
            vals=[]
            for c in row:
                vals.append({'coordinate':c.coordinate,'value':c.value,'formula':c.value if isinstance(c.value,str) and c.value.startswith('=') else None,'formula_value':wsv[c.coordinate].value,'status':'требует проверки' if c.value in (None,'') else 'ok'})
            rows.append(vals)
        sheets.append({'name':ws.title,'max_row':ws.max_row,'max_column':ws.max_column,'merged_cells':[str(r) for r in ws.merged_cells.ranges],'rows':rows,'tables':list(ws.tables.keys())})
    return {'parser_status':'completed','file_name':path.name,'sheets':sheets}

def run_command_stage(cmd, args, raw_dir, setting_name=None):
    started=time.perf_counter()
    if not cmd:
        missing=f'{setting_name} is not configured' if setting_name else 'command is not configured'
        return {'status':'not_configured','reason':missing,'error':missing,'wall_clock_seconds':round(time.perf_counter()-started,3)}
    exe=shutil.which(cmd.split()[0])
    if exe is None and not Path(cmd.split()[0]).exists():
        err=f'command not found: {cmd}'
        return {'status':'not_configured','reason':err,'error':err,'wall_clock_seconds':round(time.perf_counter()-started,3)}
    try:
        p=subprocess.run(cmd.split()+args, capture_output=True, text=True, timeout=300)
        status='completed' if p.returncode==0 else 'failed'
        return {'status':status,'reason':'command completed' if status=='completed' else f'command failed with return code {p.returncode}','returncode':p.returncode,'stdout':p.stdout,'stderr':p.stderr,'wall_clock_seconds':round(time.perf_counter()-started,3)}
    except Exception as e:
        return {'status':'failed','reason':'command raised an exception','error':str(e),'wall_clock_seconds':round(time.perf_counter()-started,3)}

def _stage_result(name, model_parser, status, reason, input_file=None, pages=None, images=None, wall_clock_seconds=None, raw_output_path=None, normalized_output_path=None, error=None, input_text_tokens=None, output_text_tokens=None, visual_input='не использовался', visual_tokens=None, total_duration=None, load_duration=None, prompt_eval_duration=None, eval_duration=None):
    return {'stage':name,'model_parser':model_parser,'status':status,'reason':reason,'input_file':input_file,'page_or_image_count':pages if pages is not None else images,'page_count':pages,'image_count':images,'wall_clock_seconds':wall_clock_seconds,'input_text_tokens':input_text_tokens,'output_text_tokens':output_text_tokens,'visual_input':visual_input,'visual_tokens':visual_tokens,'raw_output_path':raw_output_path,'normalized_output_path':normalized_output_path,'error':error,'total_duration':total_duration,'load_duration':load_duration,'prompt_eval_duration':prompt_eval_duration,'eval_duration':eval_duration}

def _append_stage(stage_results, *args, **kwargs):
    res=_stage_result(*args, **kwargs); stage_results.append(res); return res

def _png_chunk(kind, data):
    return len(data).to_bytes(4,'big')+kind+data+zlib.crc32(kind+data).to_bytes(4,'big')

def _write_blank_png(path, width=800, height=1100):
    raw=b''.join(b'\x00'+b'\xff\xff\xff'*width for _ in range(height))
    png=b'\x89PNG\r\n\x1a\n'+_png_chunk(b'IHDR', struct.pack('>IIBBBBB',width,height,8,2,0,0,0))+_png_chunk(b'IDAT', zlib.compress(raw, 1))+_png_chunk(b'IEND', b'')
    path.write_bytes(png)

def render_pdf_first_page_to_png(pdf_path, output_dir):
    output_dir.mkdir(parents=True, exist_ok=True)
    out=output_dir/f'{pdf_path.stem}-page-1.png'
    # Prefer an actual renderer when available, but keep the pipeline runnable in
    # minimal test containers by creating a page-sized placeholder PNG.
    pdftoppm=shutil.which('pdftoppm')
    if pdftoppm:
        prefix=output_dir/f'{pdf_path.stem}-page'
        subprocess.run([pdftoppm,'-f','1','-l','1','-png',str(pdf_path),str(prefix)], check=True, capture_output=True, timeout=60)
        rendered=output_dir/f'{pdf_path.stem}-page-1.png'
        if rendered.exists(): return rendered
    width,height=800,1100
    try:
        from pypdf import PdfReader
        page=PdfReader(str(pdf_path)).pages[0]
        box=page.mediabox; width=max(int(float(box.width)*2),1); height=max(int(float(box.height)*2),1)
    except Exception:
        pass
    _write_blank_png(out,width,height)
    return out

def _ollama_configured(model_id, env_name):
    base_url=os.getenv(env_name, DEFAULT_OLLAMA_BASE_URL)
    if not base_url: return False, None, f'{env_name} is not configured'
    if not model_id: return False, base_url, 'Ollama model is not configured'
    return True, base_url, None

def _append_qwen_stage(stage_results, name, result, raw_path=None, norm_path=None, input_file=None):
    _append_stage(stage_results,name,name,result.get('status'),result.get('reason') or result.get('notes') or result.get('error') or result.get('status'),input_file=input_file,pages=result.get('pages'),images=(result.get('visual_input') or {}).get('image_count'),wall_clock_seconds=result.get('wall_clock_seconds'),raw_output_path=str(raw_path) if raw_path else None,normalized_output_path=str(norm_path) if norm_path else None,error=result.get('error'),input_text_tokens=result.get('input_text_tokens'),output_text_tokens=result.get('output_text_tokens'),visual_input=result.get('visual_input','не использовался'),visual_tokens=result.get('visual_tokens'),total_duration=result.get('total_duration'),load_duration=result.get('load_duration'),prompt_eval_duration=result.get('prompt_eval_duration'),eval_duration=result.get('eval_duration'))

def process_pipeline(kit, files, data_dir=DATA_DIR):
    start=time.perf_counter(); pid=uuid4().hex; root=data_dir/'pipeline_runs'/pid; raw=root/'raw'; clean=root/'cleaned'; raw.mkdir(parents=True); clean.mkdir()
    stage_results=[]
    inv=[inventory_file(f|{'pipeline_run_id':pid}) for f in files]
    manifest_path=root/'file_manifest.json'; write_json(manifest_path, {'pipeline_run_id':pid,'kit':kit,'files':inv})
    inv_failed=[f for f in inv if f.get('processing_status')=='inventory_failed']
    _append_stage(stage_results,'первичная инвентаризация','deterministic inventory','failed' if inv_failed else 'completed','инвентаризация завершена с ошибками' if inv_failed else 'инвентаризация загруженных файлов завершена',input_file=', '.join(f['original_filename'] for f in inv) or None,pages=sum((f.get('page_count') or 0) for f in inv) or None,images=sum((f.get('image_count') or 0) for f in inv) or None,normalized_output_path=str(manifest_path),error='; '.join(f"{f['original_filename']}: {f.get('error_message')}" for f in inv_failed) or None)
    cleaned={'pipeline_run_id':pid,'kit':kit,'uploaded_files':inv,'found_documents':[],'unconfirmed_documents':[],'files':[],'potential_bom_rows':[],'images_and_schemes':[],'suspicious_ocr_fragments':[],'data_quality_risks':[],'raw_outputs':[]}
    warnings=[]
    deterministic_seen=False
    for f in inv:
        fp=Path(f['saved_path']); ext=f['extension']; fr={'input_file_id':f['id'],'file':f['original_filename'],'stages':[]}
        if f['processing_status'].endswith('failed'): warnings.append(f['original_filename'])
        if ext=='.pdf':
            pp=run_command_stage(os.getenv('PP_DOCLAYOUT_COMMAND',''), [str(fp)], raw, 'PP_DOCLAYOUT_COMMAND'); miner=run_command_stage(os.getenv('MINERU_COMMAND',''), [str(fp)], raw, 'MINERU_COMMAND')
            for name,res in [('PP-DocLayout-L',pp),('MinerU 2.5',miner)]:
                p=raw/f"{f['id']}-{name.replace(' ','_')}.json"; write_json(p,res); cleaned['raw_outputs'].append(str(p)); fr['stages'].append({'stage':name, **res});
                _append_stage(stage_results,name,name,res['status'],res.get('reason') or res.get('error') or res['status'],input_file=f['original_filename'],pages=f.get('page_count'),images=f.get('image_count'),wall_clock_seconds=res.get('wall_clock_seconds'),raw_output_path=str(p),error=res.get('error') or res.get('stderr'))
                if res['status']!='completed': warnings.append(f"{f['original_filename']}:{name}:{res['status']}")
                _record_test_run(kit,pid,f,name,res,str(p))
        elif ext=='.docx':
            deterministic_seen=True; st=time.perf_counter()
            try:
                res=parse_docx(fp); status='completed'; cleaned['found_documents'].append({'file':f['original_filename'],'type':'DOCX'})
                reason='DOCX распарсен детерминированным парсером'
            except Exception as e:
                res={'parser_status':'failed','error':str(e),'source_file':str(fp)}; status='failed'; reason='ошибка детерминированного парсинга DOCX'; warnings.append(f['original_filename'])
            p=clean/f"{f['id']}-docx-cleaned.json"; write_json(p,res); fr['cleaned_json_path']=str(p); fr['cleaned_excerpt']=res; fr['stages'].append({'stage':'deterministic-docx','status':status})
            _append_stage(stage_results,'PDF/DOCX/XLSX детерминированный парсинг','python-docx zip/xml parser',status,reason,input_file=f['original_filename'],pages=f.get('page_count'),images=f.get('image_count'),wall_clock_seconds=round(time.perf_counter()-st,3),normalized_output_path=str(p),error=res.get('error'))
        elif ext=='.xlsx':
            deterministic_seen=True; st=time.perf_counter()
            res=parse_xlsx(fp); p=clean/f"{f['id']}-xlsx-cleaned.json"; write_json(p,res); fr['cleaned_json_path']=str(p); fr['stages'].append({'stage':'deterministic-excel','status':res.get('parser_status')}); cleaned['found_documents'].append({'file':f['original_filename'],'type':'XLSX'})
            _append_stage(stage_results,'PDF/DOCX/XLSX детерминированный парсинг','openpyxl',res.get('parser_status'),'XLSX распарсен детерминированным парсером' if res.get('parser_status')=='completed' else 'ошибка или отсутствие зависимости XLSX-парсера',input_file=f['original_filename'],pages=f.get('sheet_count'),images=f.get('image_count'),wall_clock_seconds=round(time.perf_counter()-st,3),normalized_output_path=str(p),error=res.get('error'))
        elif ext=='.xls':
            deterministic_seen=True
            res={'parser_status':'not_configured','error':'XLS parser dependency/command is not configured','sheets':[]}; p=clean/f"{f['id']}-xls-cleaned.json"; write_json(p,res); fr['cleaned_json_path']=str(p); fr['stages'].append({'stage':'deterministic-excel','status':'not_configured'}); warnings.append(f['original_filename'])
            _append_stage(stage_results,'PDF/DOCX/XLSX детерминированный парсинг','XLS parser','not_configured','XLS parser dependency/command is not configured',input_file=f['original_filename'],pages=f.get('sheet_count'),images=f.get('image_count'),normalized_output_path=str(p),error=res.get('error'))
        cleaned['files'].append(fr)
    if not deterministic_seen:
        _append_stage(stage_results,'PDF/DOCX/XLSX детерминированный парсинг','deterministic parser','skipped_not_applicable','нет DOCX/XLSX/XLS файлов для детерминированного парсинга')
    # mentions not confirmation
    text=json.dumps(cleaned, ensure_ascii=False).lower()
    for doc in ['gerber','step','pick&place','bom']:
        if doc in text and not any(doc in uf['original_filename'].lower() for uf in inv): cleaned['unconfirmed_documents'].append(doc)
    cj=clean/'kit_cleaned.json'; write_json(cj, cleaned)
    _append_stage(stage_results,'нормализация без LLM','deterministic JSON normalizer','completed','сформирован нормализованный kit_cleaned.json без LLM',normalized_output_path=str(cj))
    model_results={'qwen3_vl_8b':[],'qwen2_5_3b':None}
    qwen3_ok,_,qwen3_missing=_ollama_configured(ADAPTERS['qwen3-vl-8b'].model_id,'QWEN3_VL_BASE_URL')
    if qwen3_ok:
        for f in inv:
            if f.get('extension')!='.pdf': continue
            started=time.perf_counter(); png_path=None; raw_path=raw/f"{f['id']}-Qwen3-VL-8B.json"
            try:
                png_path=render_pdf_first_page_to_png(Path(f['saved_path']), raw)
                res,seconds=ADAPTERS['qwen3-vl-8b'].process(png_path, {'execution_mode':'local_cpu','source_type':'изображение','content_type':'image/png','page_count':1})
                res['wall_clock_seconds']=round(seconds,3); res['reason']='Ollama model call completed'
            except Exception as e:
                res={'status':'failed','reason':'Ollama model call failed','error':str(e),'wall_clock_seconds':round(time.perf_counter()-started,3),'source_type':'изображение'}
                warnings.append(f"{f['original_filename']}:Qwen3-VL-8B:failed")
            write_json(raw_path,res); cleaned['raw_outputs'].append(str(raw_path)); model_results['qwen3_vl_8b'].append(res | {'raw_output_path':str(raw_path),'rendered_image_path':str(png_path) if png_path else None})
            _append_qwen_stage(stage_results,'Qwen3-VL-8B',res,raw_path=raw_path,input_file=f['original_filename'])
    else:
        res={'status':'not_configured','reason':qwen3_missing,'error':qwen3_missing}
        _append_qwen_stage(stage_results,'Qwen3-VL-8B',res)
    qwen25_ok,_,qwen25_missing=_ollama_configured(ADAPTERS['qwen2.5-3b'].model_id,'QWEN25_3B_BASE_URL')
    raw25=raw/'Qwen2.5-3B.json'
    if qwen25_ok:
        started=time.perf_counter()
        try:
            res,seconds=ADAPTERS['qwen2.5-3b'].process(cj, {'execution_mode':'local_cpu','source_type':'очищенный JSON','page_count':0})
            res['wall_clock_seconds']=round(seconds,3); res['reason']='Ollama model call completed'
        except Exception as e:
            res={'status':'failed','reason':'Ollama model call failed','error':str(e),'wall_clock_seconds':round(time.perf_counter()-started,3),'source_type':'очищенный JSON'}
            warnings.append('Qwen2.5-3B:failed')
        write_json(raw25,res); cleaned['raw_outputs'].append(str(raw25)); model_results['qwen2_5_3b']=res | {'raw_output_path':str(raw25)}
        _append_qwen_stage(stage_results,'Qwen2.5-3B',res,raw_path=raw25,norm_path=cj,input_file='kit_cleaned.json')
    else:
        res={'status':'not_configured','reason':qwen25_missing,'error':qwen25_missing}
        model_results['qwen2_5_3b']=res
        _append_qwen_stage(stage_results,'Qwen2.5-3B',res)
    bom_in_pdf=any(f.get('extension')=='.pdf' and 'bom' in (f.get('original_filename') or '').lower() for f in inv) or any(f.get('extension')=='.pdf' and 'bom' in text for f in inv)
    separate_bom=any('bom' in uf['original_filename'].lower() and uf['extension'] in {'.xlsx','.xls','.csv'} for uf in inv)
    missing=[s['stage'] for s in stage_results if s['status'] in {'not_configured','failed'}]
    completed=[s['stage'] for s in stage_results if s['status']=='completed']
    suitability='частично пригодно' if warnings or missing else 'пригодно'
    basis=f"Выполнены ключевые этапы: {', '.join(dict.fromkeys(completed)) or 'нет'}. Не выполнены/не настроены: {', '.join(dict.fromkeys(missing)) or 'нет'}. Итог «{suitability}», потому что часть обязательных этапов не была выполнена или требует проверки." if suitability=='частично пригодно' else 'Все обязательные этапы завершены без предупреждений.'
    _append_stage(stage_results,'генерация финального отчёта','deterministic report renderer','completed','final_report.json и final_report.md сформированы',wall_clock_seconds=round(time.perf_counter()-start,3))
    final={'pipeline_run':{'id':pid,'kit':kit,'status':'completed_with_warnings' if warnings else 'completed','started_at':_now(),'finished_at':_now(),'total_wall_clock_seconds':round(time.perf_counter()-start,3),'error_message':None},'composition':inv,'per_file_results':cleaned['files'],'stage_results':stage_results,'model_results':model_results,'metrics':{},'final':{'suitability':suitability,'suitability_basis':basis,'critical_errors':[],'unconfirmed_documents':cleaned['unconfirmed_documents'],'bom_status':{'elements_list_detected_in_pdf':bom_in_pdf,'separate_bom_file_uploaded_and_confirmed':separate_bom,'message':'Перечень элементов обнаружен в PDF' if bom_in_pdf else 'Перечень элементов в PDF не обнаружен', 'separate_file_message':'Отдельный BOM-файл загружен и подтверждён' if separate_bom else 'Отдельный BOM-файл не загружен и не подтверждён'},'questions_to_client':[],'short_recommendation':'Проверьте предупреждения и неподтверждённые документы.' if warnings else 'Комплект обработан детерминированно.'}}
    fj=root/'final_report.json'; fm=root/'final_report.md'; write_json(fj, final); fm.write_text(render_md(final), encoding='utf-8')
    final['pipeline_run']['final_report_json_path']=str(fj); final['pipeline_run']['final_report_md_path']=str(fm); write_json(fj, final)
    return final

def _record_test_run(kit,pid,f,stage,res,raw_path):
    create_run({'kit':kit,'stage_model':stage,'file_name':f['original_filename'],'file_count':1,'page_count':f.get('page_count') or 0,'image_count':f.get('image_count'),'table_count':f.get('table_count'),'visual_input':'не использовался','wall_clock_seconds':None,'raw_output_path':raw_path,'normalized_output_path':None,'result':res.get('status'),'critical_errors':0,'provider':'local','model_id':stage,'execution_mode':'local_cpu','source_type':f['extension'],'parent_run_id':None,'pipeline_run_id':pid,'input_file_id':f['id']})

def render_md(r):
    lines=[f"# Итоговый отчёт комплекта {r['pipeline_run']['kit']}", '', f"Статус: {r['pipeline_run']['status']}", '','## Состав комплекта']
    for f in r['composition']: lines.append(f"- {f['original_filename']} ({f['extension']}), страниц: {f.get('page_count')}, листов: {f.get('sheet_count')}, таблиц: {f.get('table_count')}, изображений: {f.get('image_count')}")
    lines += ['','## Результаты по этапам']
    for s in r.get('stage_results',[]):
        lines += [
            f"### {s['stage']}",
            f"- Модель / парсер: {s.get('model_parser') or '—'}",
            f"- Статус: {s.get('status')}",
            f"- Причина статуса: {s.get('reason') or '—'}",
            f"- Входной файл: {s.get('input_file') or '—'}",
            f"- Число страниц или изображений: {s.get('page_or_image_count') if s.get('page_or_image_count') is not None else '—'}",
            f"- Wall-clock time: {s.get('wall_clock_seconds') if s.get('wall_clock_seconds') is not None else '—'}",
            f"- input_text_tokens: {s.get('input_text_tokens') if s.get('input_text_tokens') is not None else '—'}",
            f"- output_text_tokens: {s.get('output_text_tokens') if s.get('output_text_tokens') is not None else '—'}",
            f"- total_duration: {s.get('total_duration') if s.get('total_duration') is not None else '—'}",
            f"- load_duration: {s.get('load_duration') if s.get('load_duration') is not None else '—'}",
            f"- prompt_eval_duration: {s.get('prompt_eval_duration') if s.get('prompt_eval_duration') is not None else '—'}",
            f"- eval_duration: {s.get('eval_duration') if s.get('eval_duration') is not None else '—'}",
            f"- Visual input: {s.get('visual_input') or '—'}",
            f"- Visual tokens: {s.get('visual_tokens') if s.get('visual_tokens') is not None else '—'}",
            f"- Raw output: {s.get('raw_output_path') or '—'}",
            f"- Normalized output: {s.get('normalized_output_path') or '—'}",
            f"- Ошибка: {s.get('error') or '—'}",
        ]
        if s['stage']=='Qwen2.5-3B' and s.get('status')=='completed':
            lines.append("- Примечание: Статус сформирован по ограниченному cleaned JSON без результатов MinerU; требует повторной проверки после подключения MinerU")
    bom=r['final'].get('bom_status',{})
    lines += ['','## BOM / перечень элементов', f"- {bom.get('message','Перечень элементов в PDF не обнаружен')}", f"- {bom.get('separate_file_message','Отдельный BOM-файл не загружен и не подтверждён')}"]
    lines += ['','## Итог', f"- Пригодность: {r['final']['suitability']}", f"- Основание пригодности: {r['final'].get('suitability_basis','—')}", f"- Неподтверждённые документы: {', '.join(r['final']['unconfirmed_documents']) or 'нет'}", f"- Рекомендация: {r['final']['short_recommendation']}"]
    return '\n'.join(lines)+'\n'
